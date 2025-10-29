import numpy as np
import soundfile as sf
from scipy.signal import resample
from pathlib import Path
import time
from mir_eval.separation import bss_eval_sources
from pystoi import stoi
from pesq import pesq
import pandas as pd
from openpyxl import load_workbook
import os
import museval


# Caminhos das pastas
corrupted_dir = Path("data/corrupted_dataset/")
lms_output_dir = Path("data/result_lms/")
nlms_output_dir = Path("data/result_nlms/")
klms_output_dir = Path("data/result_klms/")
nklms_output_dir = Path("data/result_nklms/")

# === Escolher apenas UM arquivo limpo ===
clean_path = Path("data/audio_limpo.wav")

tempo_execucao = []
metrics = []
output_result = Path("resultadosLMS.xlsx")

mus_lms = [ 0.6 ]
orders_lms = [ 128 ]

mus_nlms = [ 0.4 ]
orders_nlms = [ 256 ]

mus_klms = [ 0.6 ]
sigmas_klms = [ 1.4, 5 ]
dict_sizes = [ 8 ]

mus_nklms = [ 0.4 ]
sigmas_nklms = [ 1.4, 20 ]
dict_sizes_nklms = [ 8]

def snr(reference, signal):
    noise = reference - signal
    return 10 * np.log10(np.sum(reference**2) / np.sum(noise**2))

def salvar_metricas_excel(metrics, arquivo_excel="resultados.xlsx", nome_planilha="Métricas"):
    """
    Salva 'metrics' (lista de dicts) no arquivo Excel, fazendo append sem sobrescrever.
    """
    df = pd.DataFrame(metrics)

    # Se o arquivo já existe, abra com openpyxl para descobrir a última linha usada na planilha
    if os.path.exists(arquivo_excel):
        wb = load_workbook(arquivo_excel)
        if nome_planilha in wb.sheetnames:
            startrow = wb[nome_planilha].max_row  # número de linhas já existentes
        else:
            startrow = 0

        # Agora abra o ExcelWriter em modo append — não precisamos (e não devemos) setar writer.book
        with pd.ExcelWriter(arquivo_excel, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            # header só se estiver criando a planilha (startrow == 0)
            df.to_excel(writer,
                        sheet_name=nome_planilha,
                        index=False,
                        header=(startrow == 0),
                        startrow=startrow)
    else:
        # arquivo não existe: cria um novo
        with pd.ExcelWriter(arquivo_excel, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name=nome_planilha, index=False, header=True)

def metric (filter, type, file, mu, filter_order, sigma, snr, sdr, pesq, time):
    metrics.append({
            "Filtro": filter,
            "Tipo": type,
            "Arquivo": file,
            "Taxa de adaptacao": mu,
            "Ordem do filtro": filter_order,
            "Sigma": sigma,
            "snr (dB)": snr,
            "sdr (dB_)": sdr,
            "pesq": pesq,
            "Tempo (s)": time
        })

# Funções LMS
def lms_filter(x, d, mu, filter_order):
    N = len(x)
    w = np.zeros(filter_order)
    y = np.zeros(N)
    e = np.zeros(N)
    for n in range(filter_order, N):
        x_n = x[n-filter_order:n][::-1]
        y[n] = np.dot(w, x_n)
        e[n] = d[n] - y[n]
        w += mu * e[n] * x_n
    return y

# Função NLMS
def nlms_filter(x, d, mu=0.05, filter_order=512, epsilon=1e-8):
    N = len(x)
    w = np.zeros(filter_order)
    y = np.zeros(N)
    e = np.zeros(N)
    for n in range(filter_order, N):
        x_n = x[n-filter_order:n][::-1]
        norm = np.dot(x_n, x_n) + epsilon
        y[n] = np.dot(w, x_n)
        e[n] = d[n] - y[n]
        w += 2 * (mu / norm) * e[n] * x_n
    return y

# Funções de kernel
def gaussian_kernel(x, y, sigma=1.4):
    return np.exp(-np.linalg.norm(x - y) ** 2 / (2 * sigma ** 2))

def laplacian_kernel(x, y, sigma=1.4):
    return np.exp(-np.linalg.norm(x - y) / sigma)

def polynomial_kernel(x, y, sigma=None, degree=2, const=10):
    return (np.dot(x, y) + const) ** degree

# Filtro KLMS com limitação de dicionário
def klms_filter(x, d, mu=0.10, kernel_func=gaussian_kernel, sigma=1.4, max_dict_size=16):
    filter_order = max_dict_size
    N = len(x)
    y = np.zeros(N)
    alpha = []
    dictionary = []

    for n in range(filter_order, N):
        x_n = x[n - filter_order:n][::-1]
        if len(dictionary) == 0:
            y[n] = 0
        else:
            y[n] = sum(a * kernel_func(x_n, c, sigma) for a, c in zip(alpha, dictionary))

        e = d[n] - y[n]
        #print("KMLS:", e)
        alpha.append(mu * e)
        dictionary.append(x_n)

        if len(dictionary) > max_dict_size:
            dictionary.pop(0)
            alpha.pop(0)

    return y

# Filtro NKLMS com limitação de dicionário
def nklms_filter(x, d, mu=0.6, kernel_func=gaussian_kernel, sigma=20, epsilon=1e-8, max_dict_size=16):
    filter_order = max_dict_size
    N = len(x)
    y = np.zeros(N)
    alpha = []
    dictionary = []

    for n in range(filter_order, N):
        x_n = x[n - filter_order:n][::-1]
        if len(dictionary) == 0:
            y[n] = 0
        else:
            y[n] = sum(a * kernel_func(x_n, c, sigma) for a, c in zip(alpha, dictionary))

        e = d[n] - y[n]
        #print("NKMLS:", e)
        alpha.append(mu * e / (epsilon + kernel_func(x_n, x_n, sigma)))
        dictionary.append(x_n)

        if len(dictionary) > max_dict_size:
            dictionary.pop(0)
            alpha.pop(0)

    return y

# Lista de kernels a aplicar
kernels = [
#    ("gaussian", gaussian_kernel),
#    ("laplacian", laplacian_kernel),
    ("polynomial", polynomial_kernel)
]

# Escolher 1 áudio limpo fixo
clean, sr = sf.read(clean_path)
if len(clean.shape) > 1:
    clean = clean[:, 0]  # converter para mono

# === Percorrer todos os corrompidos (já existentes) ===
for corrupted_path in corrupted_dir.rglob("*.wav"):

    corrupted, sr_corrupted = sf.read(corrupted_path)
    if len(corrupted.shape) > 1:
        corrupted = corrupted[:, 0]

    min_len = min(len(clean), len(corrupted))
    clean_proc = clean[:min_len]
    corrupted_proc = corrupted[:min_len]

    relative_path = corrupted_path.relative_to(corrupted_dir)
    
    # ================== LMS ==================
    '''
    for mu in mus_lms:
        for order in orders_lms:
            start = time.time()
            y_lms = lms_filter(corrupted_proc, clean_proc, mu, order)
            tempo = time.time() - start

            if np.max(np.abs(y_lms)) < 1e-6:
                continue

            try:
                metric("lms", "", corrupted_path, mu, order, "",
                    snr(clean_proc, y_lms),
                    bss_eval_sources(np.expand_dims(clean_proc,0), np.expand_dims(y_lms,0))[0][0],
                    pesq(sr_corrupted, clean_proc, y_lms, 'wb'),
                    tempo)
            except:
                continue

            lms_out = lms_output_dir / f"mu{mu}_ord{order}" / relative_path
            lms_out.parent.mkdir(parents=True, exist_ok=True)
            sf.write(str(lms_out), y_lms, sr)

            salvar_metricas_excel(metrics, output_result)
            metrics.clear()

    # ================== NLMS ==================
    for mu in mus_nlms:
        for order in orders_nlms:
            start = time.time()
            y_nlms = nlms_filter(corrupted_proc, clean_proc, mu=mu, filter_order=order)
            tempo = time.time() - start

            if np.max(np.abs(y_nlms)) < 1e-6:
                continue

            try:
                metric("nlms", "", corrupted_path, mu, order, "",
                    snr(clean_proc, y_nlms),
                    bss_eval_sources(np.expand_dims(clean_proc,0), np.expand_dims(y_nlms,0))[0][0],
                    pesq(sr_corrupted, clean_proc, y_nlms, 'wb'),
                    tempo)
            except:
                continue

            nlms_out = nlms_output_dir / f"mu{mu}_ord{order}" / relative_path
            nlms_out.parent.mkdir(parents=True, exist_ok=True)
            sf.write(str(nlms_out), y_nlms, sr)

            salvar_metricas_excel(metrics, output_result)
            metrics.clear()
    # ================== KLMS / NKLMS ==================
    '''
    for kernel_name, kernel_func in kernels:
        for mu in mus_klms:
            for sigma in sigmas_klms:
                for dict_size in dict_sizes:
                    start = time.time()
                    y_klms = klms_filter(corrupted_proc, clean_proc,
                                         mu=mu, kernel_func=kernel_func,
                                         sigma=sigma, max_dict_size=dict_size)
                    tempo = time.time() - start
                    
                    if np.max(np.abs(y_klms)) < 1e-6:
                        continue

                    try:
                        metric("klms", kernel_name, corrupted_path, mu, dict_size, sigma,
                            snr(clean_proc, y_klms),
                            bss_eval_sources(np.expand_dims(clean_proc,0), np.expand_dims(y_klms,0))[0][0],
                            pesq(sr_corrupted, clean_proc, y_klms, 'wb'),
                            tempo)
                    except:
                        continue

                    klms_out = klms_output_dir / kernel_name / f"mu{mu}_sig{sigma}_dict{dict_size}" / relative_path
                    klms_out.parent.mkdir(parents=True, exist_ok=True)
                    sf.write(str(klms_out), y_klms, sr)

                    salvar_metricas_excel(metrics, output_result)
                    metrics.clear()

        for mu in mus_nklms:
            for sigma in sigmas_nklms:
                for dict_size in dict_sizes_nklms:
                    start = time.time()
                    y_nklms = nklms_filter(corrupted_proc, clean_proc,
                                           mu=mu, kernel_func=kernel_func,
                                           sigma=sigma, max_dict_size=dict_size)
                    tempo = time.time() - start

                    if np.max(np.abs(y_nklms)) < 1e-6:
                        continue

                    try:
                        metric("nklms", kernel_name, corrupted_path, mu, dict_size, sigma,
                            snr(clean_proc, y_nklms),
                            bss_eval_sources(np.expand_dims(clean_proc,0), np.expand_dims(y_nklms,0))[0][0],
                            pesq(sr_corrupted, clean_proc, y_nklms, 'wb'),
                            tempo)
                    except:
                        continue

                    nklms_out = nklms_output_dir / kernel_name / f"mu{mu}_sig{sigma}_dict{dict_size}" / relative_path
                    nklms_out.parent.mkdir(parents=True, exist_ok=True)
                    sf.write(str(nklms_out), y_nklms, sr)

                    salvar_metricas_excel(metrics, output_result)
                    metrics.clear()

    #print(f"Processado: {relative_path}")
#salvar_metricas_excel(metrics, "resultados.xlsx")


