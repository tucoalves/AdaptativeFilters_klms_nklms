import numpy as np
import soundfile as sf
from scipy.signal import resample
from pathlib import Path
import time
from mir_eval.separation import bss_eval_sources
from pesq import pesq
import pandas as pd
from openpyxl import load_workbook
import os
import warnings
warnings.filterwarnings("ignore", category=FutureWarning)

# Caminhos das pastas
corrupted_dir = Path("data/corrupted_dataset/")
klms_output_dir = Path("data/result_klms/")
nklms_output_dir = Path("data/result_nklms/")

# === Escolher apenas UM arquivo limpo ===
clean_path = Path("data/audio_limpo.wav")

tempo_execucao = []
metrics = []
output_result = Path("resultados_final_polinomial.xlsx")


mus_klms = [ 0.01 ]
degree_klms = [ 2 ]
const_klms = [ 5 ]
dict_sizes = [ 8 ]

mus_nklms = [ 0.6 ]
degree_nklms = [ 3 ]
const_nklms = [ 10 ]
dict_sizes_nklms = [ 8 ]

def snr(reference, signal, type):
    if np.isnan(signal).any() or np.isinf(signal).any():
                print(f"⚠️ {type} - Valores inválidos detectados — pulando este resultado.")

    noise = reference - signal
    return 10 * np.log10((np.sum(reference**2) + 1e-10) / (np.sum(noise**2) + 1e-10))

def salvar_metricas_excel(metrics, arquivo_excel="resultados.xlsx", nome_planilha="Métricas"):
    """
    Salva 'metrics' (lista de dicts) no arquivo Excel, fazendo append sem sobrescrever.
    """
    df = pd.DataFrame(metrics)

    # Se o arquivo já existe, abra com openpyxl para descobrir a última linha usada na planilha
    if os.path.exists(arquivo_excel):
        wb = load_workbook(arquivo_excel)
        if nome_planilha in wb.sheetnames:
            startrow = wb[nome_planilha].max_row  # número de linhas já existentes
        else:
            startrow = 0

        # Agora abra o ExcelWriter em modo append — não precisamos (e não devemos) setar writer.book
        with pd.ExcelWriter(arquivo_excel, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            # header só se estiver criando a planilha (startrow == 0)
            df.to_excel(writer,
                        sheet_name=nome_planilha,
                        index=False,
                        header=(startrow == 0),
                        startrow=startrow)
    else:
        # arquivo não existe: cria um novo
        with pd.ExcelWriter(arquivo_excel, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name=nome_planilha, index=False, header=True)

def metric (filter, type, file, mu, filter_order, degree, const, snr, sdr, pesq, time):
    metrics.append({
            "Filtro": filter,
            "Tipo": type,
            "Arquivo": file,
            "Taxa de adaptacao": mu,
            "Ordem do filtro": filter_order,
            "Degree": degree,
            "Constante": const,
            "snr (dB)": snr,
            "sdr (dB_)": sdr,
            "pesq": pesq,
            "Tempo (s)": time
        })


def polynomial_kernel(x, y, degree=2, const=10):
    return (np.dot(x, y) + const) ** degree

# Filtro KLMS com limitação de dicionário
def klms_filter(x, d, mu=0.10, kernel_func=polynomial_kernel, degree=2, const=1, max_dict_size=16):
    filter_order = max_dict_size
    N = len(x)
    y = np.zeros(N)
    alpha = []
    dictionary = []

    for n in range(filter_order, N):
        x_n = x[n - filter_order:n][::-1]
        if len(dictionary) == 0:
            y[n] = 0
        else:
            y[n] = sum(a * kernel_func(x_n, c, degree, const) for a, c in zip(alpha, dictionary))

        e = d[n] - y[n]
        #print("KMLS:", e)
        alpha.append(mu * e)
        dictionary.append(x_n)

        if len(dictionary) > max_dict_size:
            dictionary.pop(0)
            alpha.pop(0)

    return y

# Filtro NKLMS com limitação de dicionário
def nklms_filter(x, d, mu=0.6, kernel_func=polynomial_kernel, degree=2, const=1, epsilon=1e-8, max_dict_size=16):
    filter_order = max_dict_size
    N = len(x)
    y = np.zeros(N)
    alpha = []
    dictionary = []

    for n in range(filter_order, N):
        x_n = x[n - filter_order:n][::-1]
        if len(dictionary) == 0:
            y[n] = 0
        else:
            y[n] = sum(a * kernel_func(x_n, c, degree, const) for a, c in zip(alpha, dictionary))

        e = d[n] - y[n]
        #print("NKMLS:", e)
        alpha.append(mu * e / (epsilon + kernel_func(x_n, x_n, degree, const)))
        dictionary.append(x_n)

        if len(dictionary) > max_dict_size:
            dictionary.pop(0)
            alpha.pop(0)

    return y

# Lista de kernels a aplicar
kernels = [
    ("polynomial", polynomial_kernel)
]

# Escolher 1 áudio limpo fixo
clean, sr = sf.read(clean_path)
if len(clean.shape) > 1:
    clean = clean[:, 0]  # converter para mono


# === Percorrer todos os corrompidos (já existentes) ===
for corrupted_path in corrupted_dir.rglob("*.wav"):

    corrupted, sr_corrupted = sf.read(corrupted_path)
    if len(corrupted.shape) > 1:
        corrupted = corrupted[:, 0]
    

    min_len = min(len(clean), len(corrupted))
    clean_proc = clean[:min_len]
    corrupted_proc = corrupted[:min_len]

    relative_path = corrupted_path.relative_to(corrupted_dir)
    
    # ================== KLMS / NKLMS ==================
    
    for const in const_klms:
        for mu in mus_klms:
            for degree in degree_klms:
                for dict_size in dict_sizes:
                    start = time.time()
                    y_klms = klms_filter(corrupted_proc, clean_proc,
                                         mu=mu, kernel_func=polynomial_kernel,
                                         degree=degree, const=const, max_dict_size=dict_size)
                    tempo = time.time() - start
                    

                    try:
                        metric("klms", "Polynomial", corrupted_path, mu, dict_size, degree, const,
                            snr(clean_proc, y_klms, "KLMS-Polynomial"),
                            bss_eval_sources(np.expand_dims(clean_proc,0), np.expand_dims(y_klms,0))[0][0],
                            pesq(sr_corrupted, clean_proc, y_klms, 'wb'),
                            tempo)
                    except:
                        continue

                    klms_out = klms_output_dir / "polynomial" / f"mu{mu}_degree{degree}_dict{dict_size}_const{const}" / relative_path
                    klms_out.parent.mkdir(parents=True, exist_ok=True)
                    sf.write(str(klms_out), y_klms, sr)

                    salvar_metricas_excel(metrics, output_result)
                    metrics.clear()

    for const in const_nklms:
        for mu in mus_nklms:
            for degree in degree_nklms:
                for dict_size in dict_sizes_nklms:
                    start = time.time()
                    y_nklms = nklms_filter(corrupted_proc, clean_proc,
                                           mu=mu, kernel_func=polynomial_kernel,
                                           degree=degree, const=const, max_dict_size=dict_size)
                    tempo = time.time() - start

                    try:
                        metric("nklms", "Polynomial", corrupted_path, mu, dict_size, degree, const,
                            snr(clean_proc, y_nklms, "NKLMS-polynomial"),
                            bss_eval_sources(np.expand_dims(clean_proc,0), np.expand_dims(y_nklms,0))[0][0],
                            pesq(sr_corrupted, clean_proc, y_nklms, 'wb'),
                            tempo)
                    except:
                        continue

                    nklms_out = nklms_output_dir / "polynomial" / f"mu{mu}_degree{degree}_dict{dict_size}_const{const}" / relative_path
                    nklms_out.parent.mkdir(parents=True, exist_ok=True)
                    sf.write(str(nklms_out), y_nklms, sr)

                    salvar_metricas_excel(metrics, output_result)
                    metrics.clear()

    #print(f"Processado: {relative_path}")


